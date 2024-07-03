import openpyxl
import sympy as sp
import math
import numpy as np
from sympy import *
import pandas as pd
from scipy import integrate
from scipy.stats import norm



data_city = pd.read_csv('E:\\Undergraduate\\TD_CNBH\\table\\Example.csv') #input
list_cities = data_city['city'].values.tolist()
list_a = data_city['a'].values.tolist()
list_b = data_city['b'].values.tolist()
list_c = data_city['c'].values.tolist()
list_d = data_city['d'].values.tolist()


n = 100000
e = (1 + 1 / n) ** n

wb = openpyxl.Workbook()
ws = wb.create_sheet("result")
ws.cell(row= 1, column=1).value = 'city'
ws.cell(row= 1, column=2).value = 'a'
ws.cell(row= 1, column=3).value = 'b'
ws.cell(row= 1, column=4).value = 'd'
ws.cell(row= 1, column=5).value = 'UBD'
ws.cell(row= 1, column=6).value = 'CD'
ws.cell(row= 1, column=7).value = 'BV'
ws.cell(row= 1, column=8).value = 'ABH'
ws.cell(row= 1, column=9).value = 'TDC'


for k in range(len(list_cities)):
    filepath = list_cities[k]
    para_a = list_a[k]
    para_b = list_b[k]
    para_c = list_c[k]
    para_d = list_d[k]
    ws.cell(row=k + 2, column=1).value = filepath
    ws.cell(row=k + 2, column=2).value = para_a
    ws.cell(row=k + 2, column=3).value = para_b
    ws.cell(row=k + 2, column=4).value = para_d

    #UBD
    UBD = para_b+(2*para_c)
    ws.cell(row=k + 2, column=5).value = UBD

    #CD
    CD = ((para_a-para_d)*(e**2-1))/(2*para_c*e**2)
    ws.cell(row=k + 2, column=6).value = CD

    #BV
    V_1 = math.pi*((para_b + 2*para_c)**2) * ((para_a-para_d)/(e**2)+para_d)
    F_a = norm.cdf(2, 0, 1)
    F_b = norm.cdf(-(para_b/para_c), 0, 1)
    left_ = 2 * math.pi * para_c * para_c*(e**(-(para_b**2)/(2*para_c*para_c)) - e**(-2))
    right_ = ((2*math.pi)**1.5) * para_b *para_c *(F_a-F_b)
    V_2 = (para_a - para_d)*(left_ + right_)
    BV = V_1 + V_2
    ws.cell(row=k + 2, column=7).value = BV

    #ABH
    ABH = BV/(math.pi*UBD*UBD)
    ws.cell(row=k + 2, column=8).value = ABH

    #TDC
    rr = UBD * 1000
    para_b = list_b[k] * 1000
    para_c = list_c[k] * 1000
    BV = BV * 1000000
    min_y = (para_a-para_d) * e ** ((-(rr - para_b)**2) /(2*(para_c**2) )) +para_d
    max_y = para_a
    def f(y, a, b, c ,d):
        return 2 * math.pi * (c * (2* sp.log((a-d)/(y-d))) ** 0.5 + b) * (1 + (-c/((y-d)*(2*sp.log((a-d)/(y-d)))**0.5)) ** 2) ** 0.5
    S_1, err = integrate.quad(f, min_y, max_y, args = (para_a, para_b, para_c, para_d), limit=100)
    S_2 = 2 * math.pi * (para_b + 2 * para_c) * ((para_a - para_d) * e**(-2) + para_d)
    S_tol = S_1 + S_2
    TDC = 100 * 3 * (2 * math.pi)**2 * BV/S_tol**(3/2)
    ws.cell(row=k + 2, column=9).value = float(TDC)



    print(filepath)
wb.save(f'E:\\Undergraduate\\TD_CNBH\\Result\\Index.xlsx') #output





